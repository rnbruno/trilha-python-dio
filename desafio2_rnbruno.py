import os
import openpyxl

def menu():
    menu = """

    [cu] Criar Usuário
    [lu] Logar Usuário
    [lisu] Listar Usuários
    [s] Sair

    => """
    return input(menu)
def login():
    texto_email = """

    Insira email

    => """
    email = input(texto_email)
    texto_senha = """

    Insira senha

    => """
    senha = input(texto_senha)
    ler1 = openpyxl.load_workbook('usuarios.xlsx')
    sheetler1 = ler1.active


    for row in sheetler1.iter_rows(min_row=2):
        coluna_1, coluna_2, coluna_3, coluna_A, coluna_B = row[0], row[1], row[2], row[4], row[5]  # A primeira e segunda coluna correspondem às colunas A e B, respectivamente

        if coluna_A.value == email and coluna_B.value == senha:
            print(f"Usuário logado com sucesso.")
            logado(coluna_1.value, coluna_2.value, coluna_3.value, coluna_A.value, coluna_B.value)
            return True  # As duas strings estão na mesma linha.
    print(f"Erro ao carregar usuario.")
    return False  # As duas strings não estão na mesma linha

def verificar_usuario(email, senha):
    for row in sheet.rows:
        for cell in row:
            print(cell.value, end=" ")
        print()

def sacar(saldo, valor, extrato, numero_saques):
    saldo -= valor
    extrato += f"Saque: R$ {valor:.2f}\n"
    numero_saques += 1
    return saldo, extrato, numero_saques
def depositar(saldo, valor, extrato):
    saldo += valor
    extrato += f"Depósito: R$ {valor:.2f}\n"
    return saldo,extrato
def extratofuncao(saldo, extrato):
    print("\n================ EXTRATO ================")
    print("Não foram realizadas movimentações." if not extrato else extrato)
    print(f"\nSaldo: R$ {saldo:.2f}")
    print("==========================================")
   
def carregar_arquivo():
    usuarios = 'usuarios.xlsx'
    contas = 'contas.xlsx'
    saldo = 'saldos.xlsx'
    extrato = 'extrato.xlsx'
    if os.path.exists(usuarios) and os.path.exists(contas) and os.path.exists(saldo) and os.path.exists(extrato):
        # Se o arquivo existe, vamos ler o seu conteúdo
        ler1 = openpyxl.load_workbook(usuarios)
        sheetler1 = ler1.active
        ler2 = openpyxl.load_workbook(contas)
        sheetler2 = ler2.active
        ler3 = openpyxl.load_workbook(saldo)
        sheetler3 = ler3.active
        ler4 = openpyxl.load_workbook(extrato)
        sheetler4= ler4.active
        # Aqui você pode fazer qualquer operação que desejar com o arquivo existente
        print("Arquivo Excel encontrado. Lendo conteúdo...")
        return(sheetler1,sheetler2,sheetler3,sheetler4)
    else:
        print("aqui")
        workbook1 = openpyxl.Workbook()
        # Selecionando a primeira planilha (por padrão já existe uma)
        sheet1 = workbook1.active

        # Adicionando dados às células
        sheet1['A1'] = 'IdUsuarios'
        sheet1['B1'] = 'Nome'
        sheet1['C1'] = 'Sobrenome'
        sheet1['D1'] = 'Idade'
        sheet1['E1'] = 'Email'
        sheet1['F1'] = 'Senha'

        # Salvando o arquivo
        workbook1.save('usuarios.xlsx')
        workbook2 = openpyxl.Workbook()
        # Selecionando a primeira planilha (por padrão já existe uma)
        sheet2 = workbook2.active
        sheet2['A1'] = 'IdContas'
        sheet2['B1'] = 'Contas'
        sheet2['C1'] = 'IdUsuarios'
        workbook2.save('contas.xlsx')
        
         # Salvando o arquivo
        workbook3 = openpyxl.Workbook()
        sheet3 = workbook3.active
        sheet3['A1'] = 'IdContas'
        sheet3['B1'] = 'IdUsuarios'
        sheet3['C1'] = 'Saldo'
        workbook3.save('saldos.xlsx')
        
        workbook4 = openpyxl.Workbook()
        sheet4 = workbook4.active
        sheet4['A1'] = 'IdContas'
        sheet4['B1'] = 'IdUsuarios'
        sheet4['C1'] = 'Extrato'
        workbook4.save('extrato.xlsx')

        print("Arquivo Excel criado com sucesso.")
        return(sheet1,sheet2,sheet3,sheet4)
        
def criarusuario(nome,sobrenome, idade, email, senha):
    print("\n================ EXTRATO ================")
    print("Não foram realizadas movimentações." if not extrato else extrato)
    print(f"\nSaldo: R$ {saldo:.2f}")
    print("==========================================")

def encontrar_primeira_linha_vazia(sheet, coluna):
    linha = 1
    while sheet[coluna + str(linha)].value is not None:
        linha += 1
    return linha

def cadastrar(usuarios1):
    print("\n================ CADASTRAR USUÁRIO ================")
    while True:
        texto1 = "Informe o Primeiro Nome"
        primeiro = input(texto1)
        texto2 = "Informe o Sobrenome"
        segundo = input(texto2)
        texto3 = "Informe sua idade ex.: 15 (somente números)"
        terceiro = int(input(texto3))
        texto4 = "Informe seu email"
        quarto = input(texto4)
        texto5 = "Informe senha"
        quinto = input(texto5)
        print("\n================ VERIFIQUE DADOS ================")
        print("\nNOME: "+ primeiro)
        print("\nSOBRENOME: "+ segundo)
        print("\IDADE: "+ str(terceiro))
        print("\EMAIL: "+ quarto)
        print("\SENHA: "+ len(quinto)*"*")
        print("\nEstá correto aperte OK")
        print("\nEstá errado aperte N")
        print("\nEstá errado aperte S para sair")
        resultado =  input()
        if resultado == "OK":
            ler1 = openpyxl.load_workbook('usuarios.xlsx')
            sheetler1 = ler1.active
            primeira_linha_vazia = encontrar_primeira_linha_vazia(sheetler1, 'A')
            # Adicionar dados na primeira linha vazia
            sheetler1['A' + str(primeira_linha_vazia)] = primeira_linha_vazia
            sheetler1['B' + str(primeira_linha_vazia)] = primeiro
            sheetler1['C' + str(primeira_linha_vazia)] = segundo
            sheetler1['D' + str(primeira_linha_vazia)] = terceiro
            sheetler1['E' + str(primeira_linha_vazia)] = quarto
            sheetler1['F' + str(primeira_linha_vazia)] = quinto
            ler1.save('usuarios.xlsx')
            
            print("\nGravado com sucesso")
            break
        elif resultado == "N":
            texto1 = "Informe o Primeiro Nome"
            primeiro = input(texto1)
            texto2 = "Informe o Sobrenome"
            segundo = input(texto2)
            texto3 = "Informe sua idade ex.: 15 (somente números)"
            terceiro = int(input(texto3))
            texto4 = "Informe seu email"
            quarto = int(input(texto4))
            texto5 = "Informe senha"
            quinto = int(input(texto5))
        elif resultado == "S":
            print("\nSAIR")
            break
        else:
            print("\nErro escolha")
            print("\n================ VERIFIQUE DADOS ================")
            print("\nNOME: "+ primeiro)
            print("\nSOBRENOME: "+ segundo)
            print("\IDATE: "+ terceiro)
            print("\EMAIL: "+ quarto)
            print("\SENHA: "+ len(quinto)*"*")
            print("\nEstá correto aperte OK")
            print("\nEstá errado aperte N")
            print("\nEstá errado aperte S para sair")
# Aqui você pode continuar com o restante do seu código, utilizando o workbook e a sheet conforme necessário
def listar_usuario():
    ler1 = openpyxl.load_workbook('usuarios.xlsx')
    sheetler1 = ler1.active
    for cell_A, cell_B in zip(sheetler1['A'], sheetler1['B']):
        print(cell_A.value, cell_B.value)
def logado(coluna_1, coluna_2, coluna_3, coluna_A, coluna_B):
    print("n\BEM VINDO: "+ coluna_2+" " +coluna_3)
    menu = """

    [d] Depositar
    [s] Sacar
    [e] Extrato
    [s] Sair

    => """
    while True:
        opcao = input(menu)
        if opcao == "d":
            valor = float(input("Informe o valor do depósito: "))
            if valor > 0:
                saldo, extrato = depositar(saldo, valor, extrato)
            else:
                print("Erro! O valor informado é inválido.")

        elif opcao == "s":
            valor = float(input("Informe o valor do saque: "))

            if valor > saldo:
                print("Erro! Você não tem saldo suficiente.")

            elif valor > limite:
                print("Erro! O valor do saque excede o limite.")

            elif numero_saques >= LIMITE_SAQUES:
                print("Erro! Número máximo de saques excedido.")

            elif valor > 0:
                saldo, extrato, numero_saques = sacar(saldo, valor, extrato, numero_saques)

            else:
                print("Erro! O valor informado é inválido.")

        elif opcao == "e":
            extratofuncao(saldo,extrato)
        elif opcao == "s":
            break
        else:
            print("Operação inválida, por favor selecione novamente a operação desejada.")

def main():
    usuarios1, contas, saldos, extratos = carregar_arquivo()
    saldo = 0
    limite = 500
    extrato = ""
    numero_saques = 0
    LIMITE_SAQUES = 3
    
    
    while True:
        log = menu()
        if log == "cu":
            cadastrar(usuarios1)
        elif log == "lu":
            login()
        elif log == "lisu":
            listar_usuario()
        elif log == "s":
            break
        else:
            print("Operação inválida, por favor selecione novamente a operação desejada.")
main()